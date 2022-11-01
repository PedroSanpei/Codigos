"""
OBJETIVO CRIAR UM PROGRAMA DE LOGIN E SENHA QUE BUSCA A INFORMAÇÃO DE UM BANCO DE DADOS
"""
import openpyxl


def login_senha():
    while True:
        # 0-LOGAR OU SE CADASTRAR
        login_page = int(input("""1- Logar, 2- Cadastar: """))
        # 1-LOGAR-SE
        if login_page == 1:
            # 1.1-IMPORTAR A BIBLIOTECA PANDA
            import pandas as pd
            # 1.2-LER O EXCEL
            planilha = pd.read_excel("C:\\WORKSPACE\\Codigos\\PYTHON\\CursoPythonUdemy\\Dados.xlsx",
                                     sheet_name='dados_user')
            # 1.3-INPUTS DE LOGIN E SENHA
            login = input('Digite seu Login: ')
            password = int(input('Digite sua Senha: '))

            # 1.4-LISTA DE CHECAGEM DE LOGIN E SENHA COM BASE NA PLANILHA
            check_login = set(list(planilha['LOGIN']))
            check_senha = set(list(planilha['SENHA']))

            # 1.5-CONDIÇÕES DE VALIDAÇÃO
            if login in check_login and password in check_senha:
                print(f'Seja Bem Vindo {login}')
                break
            elif login not in check_login and password in check_senha:
                print('Login Incorreto tente novamente')
                sair = input('Tentar Novamente?  [s]im ou [n]ão: ')
                if sair == 'n':
                    break
            elif login in check_login and password not in check_senha:
                print('Senha Incorreta')
                sair = input('Tentar Novamente?  [s]im ou [n]ão: ')
                if sair == 'n':
                    break
            else:
                print('Login e Senha incorretos')

        # 2-CADASTRAR-SE

        elif login_page == 2:
            login_cadast = input('Digite nome.sobrenome: ')
            # passw_cadast = input('Digite uma senha de até 3 digitos: ')

            from openpyxl import load_workbook
            arquivo_excel = openpyxl.load_workbook("C:\\WORKSPACE\\Codigos\\PYTHON\\CursoPythonUdemy\\Dados.xlsx")
            planilha1 = arquivo_excel.active

            max_linha = planilha1.max_row
            for i in range(1, max_linha):
                planilha1.cell = login_cadast

login_senha()

